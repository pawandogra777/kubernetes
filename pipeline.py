#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import mlflow
import logging
import pandas as pd
from PIL import Image
import time
import re
import os
import base64
import requests
from pdf2image import convert_from_path
from io import BytesIO
#######for kubernetes
import os

os.makedirs("/app/OP", exist_ok=True)
##################
logging.basicConfig(
    filename="eob_pipeline77.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    filemode="a"   # append mode
)
logging.info("Pipeline started")
# ------------------ VLLM CONFIG ---------------------
vllm_url = "http://host.minikube.internal:8000/v1/chat/completions" # for kubernetes
# vllm_url = "http://172.17.0.1:8000/v1/chat/completions" # for docker
model_name = "Qwen/Qwen3-VL-8B-Instruct-FP8"
MODEL_VERSION = "Qwen3-VL-8B-FP8-v1" 
# ----------------------------------------------------

# ------------------ USER PATHS ---------------------


BASE_DIR = os.path.dirname(os.path.abspath(__file__))

input_folder = "/lambda/nfs/NDS1/EOB/EOB_1/IP"
os.makedirs(input_folder, exist_ok=True)
# output_folder = os.path.join(BASE_DIR, "OP")
#######for kuberntes
output_folder = "/lambda/nfs/NDS1/EOB/EOB_1/OP"
os.makedirs(output_folder, exist_ok=True)


payer_list_path = os.path.join(BASE_DIR, "payer_list.txt")
field_excel_path = os.path.join(BASE_DIR, "field_excel_88.xlsx")
payer_header_excel = os.path.join(BASE_DIR, "payer_headers.xlsx")
payer_json_path = os.path.join(BASE_DIR, "payer_databse.json")


# input_folder = r"/lambda/nfs/NDS1/EOB/EOB_1/IP"
# output_folder = r"/lambda/nfs/NDS1/EOB/EOB_1/OP"
os.makedirs(output_folder, exist_ok=True)
# payer_list_path = r"/lambda/nfs/NDS1/EOB/EOB_1/payer_list.txt"
with open(payer_list_path, "r") as f:
    KNOWN_PAYERS = [line.strip().upper() for line in f if line.strip()]


# ------------------ EXCEL FIELD CONFIG ---------------------
# field_excel_path = r"/lambda/nfs/NDS1/EOB/EOB_1/field_excel_88.xlsx"
field_column_name = "FIELD_NAME"

field_df = pd.read_excel(field_excel_path)
PAGE_LEVEL_FIELDS = (
    field_df[field_column_name]
    .dropna()
    .astype(str)
    .str.strip()
    .tolist()
)
PAGE_LEVEL_FIELDS_STR = ", ".join(PAGE_LEVEL_FIELDS)
# -----------------------------------------------------------

# payer_header_excel = r"/lambda/nfs/NDS1/EOB/EOB_1/payer_headers.xlsx"

payer_df = pd.read_excel(payer_header_excel)

payer_header_map = {}

for _, row in payer_df.iterrows():
    payer = str(row["Payer"]).strip().upper()
    headers = str(row["Headers"]).strip()
    header_list = [h.strip() for h in headers.split(",") if h.strip()]
    # print("header_list",header_list)
    payer_header_map[payer] = header_list
import json

# payer_json_path = r"/lambda/nfs/NDS1/EOB/EOB_1/payer_databse.json"


with open(payer_json_path, "r") as f:
# with open( r"/lambda/nfs/NDS1/EOB/EOB_1/payer_databse.json", "r") as f:
    PAYER_DATABASE = json.load(f)
mlflow.log_param("payer_database_version", "v1")
print(PAYER_DATABASE)

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
def match_known_payer1(best_payer, best_payee):
    candidates = []

    if best_payer:
        candidates.append(best_payer.upper())

    if best_payee:
        candidates.append(best_payee.upper())

    for text in candidates:
        for payer in KNOWN_PAYERS:
            if payer in text:
                return payer  # return matched database name

    return None
def match_known_payer(best_payer):
    best_payer_upper = best_payer.upper()

    for payer in KNOWN_PAYERS:
        if payer in best_payer_upper:
            return payer   # return matched database name

    return None
def highlight_keyword_failures(excel_path):
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    wb = load_workbook(excel_path)

    for ws in wb.worksheets:
        # Read header row
        headers = {
            cell.value: idx + 1
            for idx, cell in enumerate(ws[1])
            if cell.value
        }

        # Identify *_KEYWORD columns
        keyword_columns = {
            col.replace("_KEYWORD", ""): headers[col]
            for col in headers
            if col.endswith("_KEYWORD")
        }

        # Apply red fill
        for row in range(2, ws.max_row + 1):
            for base_col, keyword_col_idx in keyword_columns.items():
                keyword_value = ws.cell(row=row, column=keyword_col_idx).value

                if keyword_value and str(keyword_value).strip().upper() == "NO":
                    if base_col in headers:
                        base_col_idx = headers[base_col]
                        ws.cell(row=row, column=base_col_idx).fill = red_fill

    wb.save(excel_path)

def is_separator_row(r):
    return all((cell.strip().replace("-", "") == "") for cell in r)

def extract_table_from_text(text):
    if not text or not text.strip():
        return pd.DataFrame()
    # print("text",text)
    lines = [ln.strip() for ln in text.split("\n") if ln.strip()]
    rows = []

    for ln in lines:
        if "|" in ln:
            parts = [x.strip() for x in ln.split("|")]
        elif "," in ln:
            parts = [x.strip() for x in ln.split(",")]
        else:
            parts = re.split(r"\s{2,}", ln)

        if len(parts) > 1:
            rows.append(parts)

    if not rows:
        return pd.DataFrame()

    rows = [r for r in rows if not is_separator_row(r)]
    # print("rows",rows)
    if not rows or len(rows[0]) == 0:
        return pd.DataFrame()

    header = rows[0]
    max_len = len(header)

    df_rows = []
    for r in rows[1:]:
        if len(r) < max_len:
            r += [""] * (max_len - len(r))
        elif len(r) > max_len:
            r = r[:max_len]
        df_rows.append(r)

    return pd.DataFrame(df_rows, columns=header)


def validate_claim_text_strict(page, claim_text):
    """
    claim_text: Pipe-separated string
                ONLY first row (headers) will be used as keywords
    page: PIL Image
    """
    lines = claim_text.strip().split("\n")
    if len(lines) < 1:
        return pd.DataFrame()

    # ONLY headers (keywords)
    headers = [h.strip() for h in lines[0].split("|")]

    # Build STRICT KEYWORD-ONLY prompt
    validation_prompt = f"""
You are a STRICT KEYWORD PRESENCE VALIDATION ENGINE for an insurance EOB page.

TASK:
For EACH field name (keyword) listed below, check ONLY whether the EXACT
field name text is explicitly visible anywhere on THIS page image.

STRICT RULES:
• Validate ONLY keyword presence, NOT values.
• Do NOT infer meaning from numbers or nearby text.
• Do NOT assume synonyms, abbreviations, or similar wording.
• Case-insensitive match is allowed, but wording must be exact.
• If the keyword text is NOT visible, return NO.

FIELDS TO CHECK:
{chr(10).join(headers)}

OUTPUT FORMAT (STRICT):
• Pipe-separated
• Header row:
{ '|'.join([h + '_KEYWORD' for h in headers]) }
• Second row:
YES or NO only, in the same order

Return ONLY the table. No explanation.
"""

    # Call VLLM
    validation_text = call_vllm(page, validation_prompt)
    print("validation_text", validation_text)

    # Convert to dataframe
    df_validation = extract_table_from_text(validation_text)
    return df_validation



def call_vllm(page, prompt):
    img_bytes = page.convert("RGB")
    buf = BytesIO()
    img_bytes.save(buf, format="JPEG")
    img_b64 = base64.b64encode(buf.getvalue()).decode()

    payload = {
        "model": model_name,
        "messages": [{
            "role": "user",
            "content": [
                {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{img_b64}"}},
                {"type": "text", "text": prompt}
            ]
        }],
        "temperature": 0,
        "max_tokens": 5000
    }

    r = requests.post(vllm_url, json=payload)
    if r.status_code != 200:
        return ""
    return r.json()["choices"][0]["message"]["content"]

def merge_page_fields(df_table, df_page):
    if df_page.empty:
        return df_table

    for col in df_page.columns:
        if col not in df_table.columns:
            df_table[col] = df_page[col].iloc[0]

    return df_table


# ================== NEW: CHECK-ONLY PROMPT ==================
CHECK_ONLY_PROMPT = """
You are extracting CHECK payment details from an insurance EOB.

TASK:
Extract ONLY the following fields IF AND ONLY IF they appear TOGETHER
in the SAME CHECK payment section on the page:
- Check Date
- Check Number
- Check Amount
- MICR
PRIMARY RULE:
- If an explicit Check Number is printed, extract it.

FALLBACK RULE (STRICT):
- If Check Number is NOT explicitly printed:
  • Look for a MICR line associated with the CHECK.
  • Extract ONLY the FIRST 7 DIGITS from the MICR line.
  • Use those 7 digits as the Check Number.
- If MICR line is NOT present or does not contain at least 7 digits,
  output "" for Check Number.

MICR IDENTIFICATION RULE (MANDATORY):

- A MICR line is ALWAYS the bottom-most printed line on the check.
- It contains ONLY digits and special MICR symbols.
- It is visually separated from other text and often printed in magnetic ink style.
- Ignore any surrounding background noise or dot patterns.
- Extract digits ONLY from the MICR line.
- Do NOT extract numbers from signatures, draft numbers, or headers.

STRICT RULES:
1. Extract values ONLY if Check Date, Check Number (or MICR fallback),
   and Check Amount ALL belong to the SAME CHECK section.
2. If ANY one of the three fields is missing or belongs to a different section,
   output "" for ALL three fields.
3. Do NOT extract EFT / ACH values.
4. Do NOT guess, infer, calculate, or combine values.
5. Copy values EXACTLY as written on the page.
6. Check Date must be in MM/DD/YYYY or YYYY-MM-DD format.
7. Check Number must be numeric only.
8. Check Amount must be numeric only.
9. MICR fallback is allowed ONLY when explicit Check Number is absent.

OUTPUT FORMAT (MANDATORY — EXACT):

- Output MUST be valid CSV.
- Header MUST be EXACTLY:
  Check Date,Check Number,Check Amount,MICR
- Output MUST contain EXACTLY:
  • ONE header row
  • ONE value row

EXAMPLE OUTPUT (CHECK NUMBER FOUND):
Check Date,Check Number,Check Amount,MICR
04/26/2024,5070817,42.24, 5070817

EXAMPLE OUTPUT (MICR FALLBACK USED):
Check Date,Check Number,Check Amount
04/26/2024,1234567,42.24

EXAMPLE OUTPUT (CHECK SECTION NOT FOUND):
Check Date,Check Number,Check Amount,MICR
"","",""

Return ONLY the CSV.
No explanations.
No extra text.
"""

# ================== 4th VLLM CALL: CLAIM-BASED DETAILS ==================



# ======================================================================
# ---------------------- PROCESS FILES ----------------------
# ======================================================================
# ---------------------- PROCESS FILES ----------------------
# ======================================================================

files = [
    f for f in os.listdir(input_folder)
    if f.lower().endswith(('.png', '.jpg', '.jpeg', '.tiff', '.tif', '.pdf'))
]

print(f"\nFound {len(files)} files\n")
mlflow.set_experiment("EOB_Extraction_Pipeline")

if mlflow.active_run():
    mlflow.end_run()
total_empty_table_pages = 0
total_payer_match_failed = 0
total_files_processed = 0
pipeline_start_time = time.time()
for file in files:
    with mlflow.start_run():

            # with mlflow.start_run():

        mlflow.log_param("model_name", model_name)
        mlflow.log_param("llm_model_version", MODEL_VERSION)
        mlflow.log_param("vllm_url", vllm_url)
        mlflow.log_param("total_input_files", len(files))


        print(f"\nProcessing: {file}")
        file_start_time = time.time()
        mlflow.log_param("current_file", file)
        logging.info(f"Processing file: {file}")
        logging.info(f"Processing file: {file}")
        logging.info(f"Model Name: {model_name}")
        logging.info(f"Model Version: {MODEL_VERSION}")
        print(f"Start time: {time.strftime('%H:%M:%S')}")

        file_path = os.path.join(input_folder, file)

        # ================== LOAD PAGES ==================
        try:
            pages = []

            if file.lower().endswith(".pdf"):
                pages = convert_from_path(file_path, dpi=300)
            else:
                img = Image.open(file_path)
                if hasattr(img, "n_frames") and img.n_frames > 1:
                    for i in range(img.n_frames):
                        img.seek(i)
                        pages.append(img.copy())
                else:
                    pages = [img]

            print("Total pages detected:", len(pages))

        except Exception as e:
            print("Open error:", e)
            continue

        # ==========================================================
        # ================== STEP 1: CHECK EXTRACTION ==============
        # ==========================================================

        df_check = pd.DataFrame(columns=["Check Date", "Check Number", "Check Amount", "MICR"])

        if pages:
            print("Extracting CHECK details from FIRST PAGE only")
            check_text = call_vllm(pages[0], CHECK_ONLY_PROMPT)
            df_tmp = extract_table_from_text(check_text)

            if not df_tmp.empty:
                df_check = df_tmp.iloc[:1]

        PAGE_PROMPT = f"""

            You are extracting FIXED, PAGE-LEVEL fields from page.

            IMPORTANT CONTEXT (MANDATORY):
            - Extract data from EACH PAGE independently.

            The page contains these extractable fields IN THIS EXACT ORDER:
            1. Payer Name
            2. Payee Name

            TASK:
            Extract EXACTLY the following fields:
            {PAGE_LEVEL_FIELDS_STR}

            FIELD DEFINITIONS (STRICT — DO NOT MIX VALUES):

            Payer Name:
            - MUST be an INSURANCE COMPANY.
            - MUST include the COMPLETE postal address immediately associated with the payer.
            - Address MUST include ALL of the following TOGETHER:
              • Street or PO Box
              • City
              • State
              • ZIP code
            - Combine payer name and full address into ONE continuous value.
            - DO NOT extract payee, hospital, provider, or patient information here.

            Payee Name:
            - MUST be a HOSPITAL, CLINIC, MEDICAL GROUP, PROVIDER, OR EMS / AMBULANCE / Rescue / RESCUE SERVICE.
            - MUST include the COMPLETE postal address immediately associated with the provider.
            - Address MUST include ALL of the following TOGETHER:
              • Street or PO Box
              • City
              • State
              • ZIP code
            - Combine provider name and full address into ONE continuous value.
            - DO NOT extract insurance or payer information here.
            - Payee MAY be an individual or provider receiving payment.


            ABSOLUTE ADDRESS REQUIREMENT (HARD STOP RULE):
            - Extract Payer Name or Payee Name ONLY IF the COMPLETE postal address is present.
            - COMPLETE address means Street/PO Box + City + State + ZIP present TOGETHER.
            - If ANY component is missing:
              • DO NOT extract that field
              • Output "" for that field
            - DO NOT extract partial addresses.
            - DO NOT extract name-only entries.
            - DO NOT combine text from different page locations.
            - Name and address MUST be visually and textually contiguous.

            LOGO–ADDRESS ASSOCIATION RULE (MANDATORY):
            - The Payer name MAY appear as a LOGO or standalone word (e.g., "DMBA").
            - If a LOGO or short name appears near the top of the page:
              • It may be associated with the nearest complete postal address
              • ONLY if that address clearly belongs to the same entity
            - The logo name and address do NOT need to be on the same line.
            - DO NOT associate a logo with a BANK address.
            - DO NOT associate across different page sections.


            BANK ROLE DISAMBIGUATION (CRITICAL):
            - A BANK may appear on the page ONLY as the ISSUING or PROCESSING institution.
            - The ISSUING BANK is NOT the Payer and NOT the Payee.
            - If multiple organizations appear:
              • Identify the entity RESPONSIBLE FOR PAYMENT as the Payer.
              • Ignore the bank even if it has a full address.
            - NEVER extract an issuing bank as Payer or Payee.

            BANK AND NUMBER EXCLUSION RULE (MANDATORY):
            - BANKS or FINANCIAL INSTITUTIONS are NEVER Payer or Payee.
            - Exclude values containing:
              Bank, Banc, Banking, Financial, Credit Union, Trust
            - If a candidate value STARTS WITH a number:
              • Treat it as NOT FOUND
              • Output ""
            ENTITY PRIORITY RULE (MANDATORY):
            - Prefer NON-BANK entities when selecting Payer or Payee.
            - If a BANK and a NON-BANK entity both have complete postal addresses:
              → ALWAYS select the NON-BANK entity.
            - Issuing or processing banks MUST be ignored even if a full address is present.

            ABSOLUTE BANK ADDRESS EXCLUSION (HARD STOP):
            - ANY address associated with a BANK or FINANCIAL INSTITUTION is INVALID.
            - This includes lockbox addresses, payment remittance addresses, or mailing addresses.
            - If the name OR address contains bank-related terms such as:
              Bank, N.A., N.A., N A, Chase, JPMorgan, Wells Fargo, Citi, Citibank, PNC, BOA, BofA
              → DO NOT extract.
            - Even if a FULL address appears, treat it as NOT FOUND.
            - Output "" for that field.

            STRICT RULES (NON-NEGOTIABLE):
            1. Extract ONLY the fields listed above.
            2. Each field MUST have EXACTLY ONE value.
            3. Full address is REQUIRED for extraction.
            4. NEVER mix payer and provider information.
            5. NEVER guess, infer, or assume.
            6. Copy text EXACTLY as written.
            7. Preserve original line order.
            8. If a field is NOT FOUND of the page, output "".

            FLATTEN MULTI-LINE TEXT (MANDATORY):
            - For each field, merge all lines into ONE single line.
            - Use exactly ONE space between lines.
            - Do NOT include any line breaks.
            - The final output MUST be a single continuous line per field.


            CRITICAL PIPE NORMALIZATION RULE (MANDATORY):
            - Output MUST contain EXACTLY 2 fields.
            - MUST NOT start or end with |.
            - Missing fields MUST be explicitly written as "".
            - Output EACH field as ONE SINGLE CONTINUOUS LINE.
            - Correct example:
              ""|Behavioral Health of North Texas LLP PO BOX 639746 CINCINNATI OH 45263-9746
              Aetna Better Health of Texas P.O. Box 982979 El Paso, TX 79998-2979|""
              Aetna Better Health of Texas P.O. Box 982979 El Paso, TX 79998-2979|Behavioral Health of North Texas LLP PO BOX 639746 CINCINNATI OH 45263-9746
            - Incorrect:
              |Aetna Better Health|
              ||Aetna Better Health

            LINE NORMALIZATION RULE:
            - Each field MUST be a SINGLE LINE.
            - Merge multi-line text using ONE space.
            - DO NOT insert line breaks.

            OUTPUT FORMAT (MANDATORY):
            - PIPE-separated output.
            - FIRST row: header.
            - SECOND row: values.
            - EXACTLY one header and one value row.
            - NO explanations, NO extra text.

            OUTPUT FORMAT EXAMPLE:
            Payer Name|Payee Name
            Aetna Better Health of Texas P.O. Box 982979 El Paso, TX 79998-2979|Behavioral Health of North Texas LLP PO BOX 639746 CINCINNATI OH 45263-9746
    """
        # ==========================================================
        # ================== STEP 2: PASS 1 (GET BEST PAYER) =======
        # ==========================================================

        best_payer = ""
        best_payee = ""

        print("\nPASS 1 → Detecting Best Payer & Payee")

        for i, page in enumerate(pages[:5]):   # only first 5 pages

            w, h = page.size
            image_for_llm = page if i == 0 else page.crop((0, 0, w, int(h * 0.25)))

            page_text = call_vllm(image_for_llm, PAGE_PROMPT)
            print("page_text",page_text)
            df_page = extract_table_from_text(page_text)

            if not df_page.empty:

                payer = df_page.get("Payer Name", [""])[0].strip()
                payee = df_page.get("Payee Name", [""])[0].strip()

                if len(payer) > len(best_payer):
                    best_payer = payer

                if len(payee) > len(best_payee):
                    best_payee = payee

        matched_payer = match_known_payer1(best_payer, best_payee)

        print("FINAL Best Payer:", best_payer)
        print("FINAL Best Payee:", best_payee)
        print("Matched Payer:", matched_payer)
        # print("PAYER_DATABASE",PAYER_DATABASE)
        if not matched_payer or matched_payer not in PAYER_DATABASE:
            total_payer_match_failed += 1   # ✅ ADD THIS
            mlflow.log_metric("payer_match_failed", 1)
            logging.warning(f"Payer not matched for file: {file}")
            continue
        # ==========================================================
        # ================== STEP 3: PASS 2 (TABLE + CLAIM) ========
        # ==========================================================

        print("\nPASS 2 → Extracting Tables & Claim Data")

        all_pages = []
        claim_pages = []

        # header_list = payer_header_map[matched_payer]
        # header_string = "\n- ".join(header_list)
        payer_example = PAYER_DATABASE[matched_payer]["example"]

        header_list = payer_header_map[matched_payer]
        print("header_list",header_list)
        header_string = "\n- ".join(header_list)


        CLAIM_BASED_PROMPT = f"""
            You are extracting table data from an insurance EOB page.

            Payer identified as: {matched_payer}

            Extract ONLY the following columns:

            - {header_string}

            STRICT RULES:
            1. Extract ONLY rows that contain these headers.
            2. Do NOT extract any other columns.
            3. Maintain exact column order as listed above.
            4. If a value is missing, output "".
            5. Do NOT guess or infer.
            6. Ignore summary tables and non-service-line sections.
            7. STOP extracting rows ONLY when a new Claim Number section starts.
            8. Include rows that contain:
               • Claim Total
               • Claim Sub-Total
               • Claim Subtotal
               • Column Total
               • Column Totals
               • Totals
               • TOTAL
               • Total lines
               • Any row that clearly represents the SUM for that Claim Number

            OUTPUT FORMAT:
            - First row: EXACT header names as listed above.
            - Subsequent rows: data.
            - Pipe-separated format only.
            - No explanations.

    STRICT RULES:
    1. Claim Number MUST be present; otherwise do not output any rows.
    2. Header MUST appear exactly once (first row only).
    3. Do NOT include explanations, labels, or commentary.
    4. Return ONLY the pipe-separated output.

    """
        TABLE_PROMPT = f"""

        You are extracting table data from an insurance EOB page.

        Payer identified as: {matched_payer}


        IMPORTANT:
        Below is an EXAMPLE of how this payer formats their table.

        EXAMPLE TABLE FORMAT:
        {payer_example}

        CRITICAL HEADER VALIDATION:

        You are given RAW OCR TEXT of this page.

        STEP 1:
        Search the OCR text for the EXACT header string below:

        "{payer_example.splitlines()[0]}"

        STEP 2:
        If this exact header string does NOT appear character-by-character in OCR:
        → RETURN NOTHING
        → DO NOT generate header
        → DO NOT reuse previous page data
        → DO NOT infer

        STEP 3:
        Only if the exact header line is found in this page's OCR text,
        then extract rows under that header.

        TASK (Only if header exists):

        - Extract ONLY that table that matches the ABOVE example structure.
        - Follow EXACT column order as shown in the example header.

        STRICT ROW CONTINUITY RULE (MANDATORY):

        1. You MUST extract EVERY SINGLE row that appears under the header.
        2. Do NOT skip ANY service line.
        3. Do NOT merge rows.
        4. Do NOT summarize repetitive CPT codes.
        5. Maintain exact top-to-bottom order.
        6. Include ALL service lines until the CLAIM TOTALS (or Totals row).
        7. If even ONE visible row is skipped → RETURN NOTHING.

        CRITICAL ROW COUNT VALIDATION:

        - Count the number of visible rows between the header and the Totals row.
        - Output MUST contain the exact same number of rows.
        - If row count does NOT match → RETURN NOTHING.

        - If a value is missing, output "".
        - Do NOT guess or infer.
        - Ignore unrelated tables.

        OUTPUT RULES:

        - First row MUST be the header (exactly as shown in example).
        - Pipe-separated format only.
        - No explanations.
        - If header not found → output NOTHING.
        """


        for i, page in enumerate(pages):

            print(f"\nProcessing Page {i+1}")

            # ---------------- TABLE EXTRACTION ----------------
            table_text = call_vllm(page, TABLE_PROMPT)
            print("table_text",table_text)
            df_table = extract_table_from_text(table_text)

            if df_table.empty:
                mlflow.log_metric("empty_table_pages", 1)
                total_empty_table_pages += 1   # ✅ ADD THIS
                logging.warning(f"Empty table on page {i+1}")
                continue

            # ---------------- CLAIM EXTRACTION ----------------
            claim_text = call_vllm(page, CLAIM_BASED_PROMPT)
            print("claim_text",claim_text)
            df_claim = extract_table_from_text(claim_text)

            if not df_claim.empty:
                df_claim = df_claim.loc[:, ~df_claim.columns.duplicated(keep='first')]
                df_claim = df_claim.loc[:, df_claim.apply(lambda col: col.astype(str).str.strip().any())]

            # df_validation = validate_claim_text_strict(page, claim_text)

            # if not df_validation.empty and not df_claim.empty:
            #     df_claim = df_claim.reset_index(drop=True)
            #     for col in df_validation.columns:
            #         df_claim[col] = df_validation[col].iloc[0]

            if df_claim.empty:
                mlflow.log_metric("empty_claim_sections", 1)
                df_claim = pd.DataFrame([{
                    "Claim Number": "",
                    "Patient Name": "",
                    "Patient Account Number": "",
                    "Patient ID": "",
                    "Member/Subscriber/Insured Name": "",
                    "Member/Subscriber/Insured ID": ""
                }])

            # ---------------- MERGE PAGE FIELDS ----------------
            df_table["Payer Name"] = best_payer
            df_table["Payee Name"] = best_payee

            # ---------------- MERGE CHECK FIELDS ----------------
            if not df_check.empty:
                for col in df_check.columns:
                    df_table[col] = df_check[col].iloc[0]
            else:
                df_table["Check Date"] = ""
                df_table["Check Number"] = ""
                df_table["Check Amount"] = ""
                df_table["MICR"] = ""

            df_table["PAGE_NO"] = i + 1

            all_pages.append((i + 1, df_table))
            claim_pages.append((i + 1, df_claim))

        # ==========================================================
        # ================== SAVE OUTPUT ============================
        # ==========================================================

        if all_pages:

            out_path = os.path.join(output_folder, f"{os.path.splitext(file)[0]}_FINAL.xlsx")
            with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                for pno, df in all_pages:
                    df.to_excel(writer, sheet_name=f"PAGE_{pno}", index=False)

            print("Saved:", out_path)
            total_rows = sum(len(df) for _, df in all_pages)
            mlflow.log_metric("total_extracted_rows", total_rows)
        if claim_pages:

            claim_out_path = os.path.join(output_folder, f"{os.path.splitext(file)[0]}_CLAIM_BASED.xlsx")
            with pd.ExcelWriter(claim_out_path, engine="openpyxl") as writer:
                for pno, dfc in claim_pages:
                    dfc.to_excel(writer, sheet_name=f"PAGE_{pno}", index=False)

            highlight_keyword_failures(claim_out_path)
            print("Saved:", claim_out_path)
            total_rows = sum(len(df) for _, df in all_pages)
            mlflow.log_metric("total_extracted_rows", total_rows)

        # ---------------- TIME LOG ----------------
        file_end_time = time.time()
        total_seconds = file_end_time - file_start_time

        minutes = int(total_seconds // 60)
        seconds = round(total_seconds % 60, 2)
        mlflow.log_metric("processing_time_seconds", total_seconds)
        mlflow.log_metric("pages_in_file", len(pages))
        logging.info(f"Completed file: {file} in {total_seconds} sec")
        total_files_processed += 1
        print(f"File completed: {file} | {minutes} min {seconds} sec")
pipeline_end_time = time.time()
total_pipeline_time = pipeline_end_time - pipeline_start_time

mlflow.log_metric("total_files_processed", total_files_processed)
mlflow.log_metric("total_empty_table_pages", total_empty_table_pages)
mlflow.log_metric("total_payer_match_failed", total_payer_match_failed)
mlflow.log_metric("total_pipeline_time_seconds", total_pipeline_time)

logging.info(f"PIPELINE SUMMARY | Files={total_files_processed} | EmptyTables={total_empty_table_pages} | PayerFails={total_payer_match_failed} | TotalTime={total_pipeline_time}")

